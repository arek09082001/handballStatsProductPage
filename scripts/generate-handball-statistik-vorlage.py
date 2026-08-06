#!/usr/bin/env python3
"""Erzeugt die kostenlose Handball-Statistik Excel-Vorlage.

Ausgabe: public/downloads/handball-statistik-vorlage.xlsx

Die Vorlage besteht aus vier Blättern:

  Kader           – Mannschaftsliste (Eingabe)
  Spielprotokoll  – eine Zeile pro Aktion (Eingabe)
  Auswertung      – Werte pro Spieler für EIN Spiel (Formeln)
  Saison          – Spielliste und rollierende Saisonwerte (Formeln)

Alle Kennzahlen werden per Formel aus dem Spielprotokoll berechnet – es gibt
keine hart eingetragenen Ergebnisse. Nach dem Erzeugen wird die Datei mit
LibreOffice neu berechnet (scripts/recalc.py der xlsx-Skill), damit
Vorschauen und Tabellen-Apps sofort Werte anzeigen.

Aufruf:  python3 scripts/generate-handball-statistik-vorlage.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = Path(__file__).resolve().parents[1] / "public" / "downloads" / "handball-statistik-vorlage.xlsx"

# --- Design: Trainertafel-Farben (siehe DESIGN.md) -------------------------
COURT = "12333D"          # dunkler Hallenboden – Kopfzeilen
MARKER = "F97316"         # Marker-Orange – Akzente
PAPER_2 = "F0EAE0"        # Papier-Panel – Eingabefelder
EXAMPLE = "E7E0D4"        # Beispielzeilen
INK = "1B2230"

FONT = "Arial"

HEAD_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT, size=14, bold=True, color=INK)
NOTE_FONT = Font(name=FONT, size=9, italic=True, color="5A6472")
BODY_FONT = Font(name=FONT, size=10, color=INK)
BOLD_FONT = Font(name=FONT, size=10, bold=True, color=INK)

HEAD_FILL = PatternFill("solid", fgColor=COURT)
INPUT_FILL = PatternFill("solid", fgColor=PAPER_2)
EXAMPLE_FILL = PatternFill("solid", fgColor=EXAMPLE)
TOTAL_FILL = PatternFill("solid", fgColor="FDE4CE")

THIN = Side(style="thin", color="C9C2B6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PERCENT = "0.0%"

# --- Datenbereiche des Spielprotokolls ------------------------------------
FIRST_ROW = 4
LAST_ROW = 403
P_SPIEL = f"Spielprotokoll!$A${FIRST_ROW}:$A${LAST_ROW}"
P_NR = f"Spielprotokoll!$C${FIRST_ROW}:$C${LAST_ROW}"
P_AKTION = f"Spielprotokoll!$E${FIRST_ROW}:$E${LAST_ROW}"

KADER_FIRST = 4
KADER_LAST = 23  # 20 Kaderplätze

AKTIONEN = [
    "Tor",
    "Fehlwurf",
    "7m Tor",
    "7m Fehlwurf",
    "Assist",
    "Techn. Fehler",
    "Parade",
    "Gegentor",
    "2 Minuten",
    "Gelbe Karte",
    "Rote Karte",
]

ZONEN = [
    "Außen",
    "Kreis",
    "Rückraum",
    "Durchbruch",
    "Tempogegenstoß",
    "Siebenmeter",
]

POSITIONEN = [
    "Torwart",
    "Linksaußen",
    "Rückraum links",
    "Rückraum Mitte",
    "Rückraum rechts",
    "Rechtsaußen",
    "Kreisläufer",
]

BEISPIEL_KADER = [
    (1, "Tim Bergmann", "Torwart", 2003),
    (7, "Lukas Sander", "Rückraum Mitte", 2004),
    (9, "Nils Hofer", "Kreisläufer", 2002),
    (11, "Jan Weber", "Linksaußen", 2005),
]

BEISPIEL_PROTOKOLL = [
    ("Sp1", 3, 7, "Tor", "Rückraum", ""),
    ("Sp1", 5, 9, "Tor", "Kreis", ""),
    ("Sp1", 7, 7, "Fehlwurf", "Rückraum", "geblockt"),
    ("Sp1", 9, 1, "Parade", "", ""),
    ("Sp1", 11, 1, "Gegentor", "", ""),
    ("Sp1", 14, 11, "Tor", "Außen", ""),
    ("Sp1", 17, 7, "Techn. Fehler", "", "Schrittfehler"),
    ("Sp1", 19, 1, "Parade", "", ""),
    ("Sp1", 22, 9, "Fehlwurf", "Kreis", ""),
    ("Sp1", 24, 7, "7m Tor", "Siebenmeter", ""),
    ("Sp1", 26, 11, "Assist", "", ""),
    ("Sp1", 28, 7, "2 Minuten", "", ""),
    ("Sp1", 31, 1, "Gegentor", "", ""),
    ("Sp1", 35, 9, "Tor", "Kreis", ""),
]


def style_header(ws, row: int, headers: list[str]) -> None:
    for index, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=index, value=title)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def set_widths(ws, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def title_block(ws, title: str, note: str) -> None:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = note
    ws["A2"].font = NOTE_FONT
    ws.row_dimensions[2].height = 26


def print_setup(ws, landscape: bool = False, repeat_row: str | None = None) -> None:
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    if repeat_row:
        ws.print_title_rows = repeat_row


def build_kader(ws) -> None:
    title_block(
        ws,
        "Kader",
        "Hier trägst du deine Mannschaft ein. Die Rückennummer ist der Schlüssel für alle anderen Blätter – "
        "sie muss eindeutig sein. Die grau hinterlegten Zeilen sind Beispiele: einfach überschreiben.",
    )
    style_header(ws, 3, ["Nr.", "Name", "Position", "Jahrgang", "Notiz"])

    for row in range(KADER_FIRST, KADER_LAST + 1):
        is_example = row - KADER_FIRST < len(BEISPIEL_KADER)
        values = BEISPIEL_KADER[row - KADER_FIRST] if is_example else ("", "", "", "")
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=column, value=value if value != "" else None)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.fill = EXAMPLE_FILL if is_example else INPUT_FILL
            cell.alignment = Alignment(horizontal="center" if column in (1, 4) else "left")
        notiz = ws.cell(row=row, column=5)
        notiz.border = BORDER
        notiz.font = BODY_FONT
        notiz.fill = EXAMPLE_FILL if is_example else INPUT_FILL

    position_validation = DataValidation(
        type="list", formula1='"' + ",".join(POSITIONEN) + '"', allow_blank=True
    )
    position_validation.error = "Bitte eine Position aus der Liste wählen."
    ws.add_data_validation(position_validation)
    position_validation.add(f"C{KADER_FIRST}:C{KADER_LAST}")

    set_widths(ws, {"A": 6, "B": 26, "C": 18, "D": 10, "E": 30})
    ws.freeze_panes = "A4"
    print_setup(ws, repeat_row="3:3")


def build_protokoll(ws) -> None:
    title_block(
        ws,
        "Spielprotokoll",
        "Eine Zeile pro Aktion – während des Spiels oder hinterher vom Videomitschnitt. "
        "Spiel und Nr. füllst du selbst, der Name kommt automatisch aus dem Kader. "
        "Aktion und Zone lassen sich per Auswahlliste setzen.",
    )
    style_header(ws, 3, ["Spiel", "Minute", "Nr.", "Name", "Aktion", "Zone", "Notiz"])

    for row in range(FIRST_ROW, LAST_ROW + 1):
        example_index = row - FIRST_ROW
        is_example = example_index < len(BEISPIEL_PROTOKOLL)
        fill = EXAMPLE_FILL if is_example else INPUT_FILL

        if is_example:
            spiel, minute, nummer, aktion, zone, notiz = BEISPIEL_PROTOKOLL[example_index]
            ws.cell(row=row, column=1, value=spiel)
            ws.cell(row=row, column=2, value=minute)
            ws.cell(row=row, column=3, value=nummer)
            ws.cell(row=row, column=5, value=aktion)
            ws.cell(row=row, column=6, value=zone or None)
            ws.cell(row=row, column=7, value=notiz or None)

        # Name wird immer aus dem Kader gezogen (Nachschlag über die Nummer).
        ws.cell(
            row=row,
            column=4,
            value=(
                f'=IF($C{row}="","",'
                f'IFERROR(INDEX(Kader!$B${KADER_FIRST}:$B${KADER_LAST},'
                f'MATCH($C{row},Kader!$A${KADER_FIRST}:$A${KADER_LAST},0)),"unbekannte Nr."))'
            ),
        )

        for column in range(1, 8):
            cell = ws.cell(row=row, column=column)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center" if column in (2, 3) else "left")

    aktion_validation = DataValidation(
        type="list", formula1='"' + ",".join(AKTIONEN) + '"', allow_blank=True
    )
    aktion_validation.error = "Bitte eine Aktion aus der Liste wählen."
    ws.add_data_validation(aktion_validation)
    aktion_validation.add(f"E{FIRST_ROW}:E{LAST_ROW}")

    zone_validation = DataValidation(
        type="list", formula1='"' + ",".join(ZONEN) + '"', allow_blank=True
    )
    ws.add_data_validation(zone_validation)
    zone_validation.add(f"F{FIRST_ROW}:F{LAST_ROW}")

    set_widths(ws, {"A": 14, "B": 8, "C": 6, "D": 24, "E": 15, "F": 17, "G": 26})
    ws.freeze_panes = "A4"
    print_setup(ws, repeat_row="3:3")


def counts(criteria: str, nr_cell: str, spiel_cell: str | None) -> str:
    """COUNTIFS über das Protokoll – optional auf ein Spiel eingegrenzt."""
    if spiel_cell:
        return f'COUNTIFS({P_SPIEL},{spiel_cell},{P_NR},{nr_cell},{P_AKTION},"{criteria}")'
    return f'COUNTIFS({P_NR},{nr_cell},{P_AKTION},"{criteria}")'


def build_auswertung(ws) -> None:
    title_block(
        ws,
        "Auswertung (ein Spiel)",
        "Trage oben die Spiel-Bezeichnung ein (dieselbe wie im Protokoll, z. B. Sp1) – alle Werte rechnen sich "
        "automatisch. Würfe = Tore + Fehlwürfe aus dem Feld; Siebenmeter zählen separat. "
        "± = Tore (inkl. 7 m) − Fehlwürfe (inkl. 7 m) − technische Fehler.",
    )

    ws["A3"] = "Spiel:"
    ws["A3"].font = BOLD_FONT
    ws["B3"] = "Sp1"
    ws["B3"].font = Font(name=FONT, size=11, bold=True, color="C2410C")
    ws["B3"].fill = INPUT_FILL
    ws["B3"].border = BORDER
    ws["B3"].alignment = Alignment(horizontal="center")

    spiel_validation = DataValidation(type="list", formula1="=Spiele", allow_blank=True)
    ws.add_data_validation(spiel_validation)
    spiel_validation.add("B3")

    headers = [
        "Nr.",
        "Name",
        "Würfe",
        "Tore",
        "Wurfquote",
        "7m Würfe",
        "7m Tore",
        "7m-Quote",
        "Assists",
        "Techn. Fehler",
        "Zeitstrafen",
        "±",
        "Paraden",
        "Gegentore",
        "Paradenquote",
    ]
    header_row = 5
    style_header(ws, header_row, headers)

    first = header_row + 1
    last = first + (KADER_LAST - KADER_FIRST)

    for offset, row in enumerate(range(first, last + 1)):
        kader_row = KADER_FIRST + offset
        nr = f"$A{row}"
        guard = f'IF($A{row}="","",'

        tore = counts("Tor", nr, "$B$3")
        fehl = counts("Fehlwurf", nr, "$B$3")
        sieben_tor = counts("7m Tor", nr, "$B$3")
        sieben_fehl = counts("7m Fehlwurf", nr, "$B$3")
        techn = counts("Techn. Fehler", nr, "$B$3")

        formulas = {
            1: f'=IF(Kader!$A{kader_row}="","",Kader!$A{kader_row})',
            2: f'=IF(Kader!$B{kader_row}="","",Kader!$B{kader_row})',
            3: f"={guard}{tore}+{fehl})",
            4: f"={guard}{tore})",
            5: f'={guard}IFERROR($D{row}/$C{row},""))',
            6: f"={guard}{sieben_tor}+{sieben_fehl})",
            7: f"={guard}{sieben_tor})",
            8: f'={guard}IFERROR($G{row}/$F{row},""))',
            9: f'={guard}{counts("Assist", nr, "$B$3")})',
            10: f"={guard}{techn})",
            11: f'={guard}{counts("2 Minuten", nr, "$B$3")})',
            12: f"={guard}($D{row}+$G{row})-($C{row}-$D{row})-($F{row}-$G{row})-$J{row})",
            13: f'={guard}{counts("Parade", nr, "$B$3")})',
            14: f'={guard}{counts("Gegentor", nr, "$B$3")})',
            15: f'={guard}IFERROR($M{row}/($M{row}+$N{row}),""))',
        }

        for column, formula in formulas.items():
            cell = ws.cell(row=row, column=column, value=formula)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if column != 2 else "left")
            if column in (5, 8, 15):
                cell.number_format = PERCENT

    total_row = last + 1
    ws.cell(row=total_row, column=1, value="")
    ws.cell(row=total_row, column=2, value="Mannschaft gesamt")
    for column in (3, 4, 6, 7, 9, 10, 11, 12, 13, 14):
        letter = get_column_letter(column)
        ws.cell(row=total_row, column=column, value=f"=SUM({letter}{first}:{letter}{last})")
    ws.cell(row=total_row, column=5, value=f'=IFERROR($D{total_row}/$C{total_row},"")')
    ws.cell(row=total_row, column=8, value=f'=IFERROR($G{total_row}/$F{total_row},"")')
    ws.cell(
        row=total_row,
        column=15,
        value=f'=IFERROR($M{total_row}/($M{total_row}+$N{total_row}),"")',
    )
    for column in range(1, 16):
        cell = ws.cell(row=total_row, column=column)
        cell.font = BOLD_FONT
        cell.fill = TOTAL_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center" if column != 2 else "left")
        if column in (5, 8, 15):
            cell.number_format = PERCENT

    ws.cell(
        row=total_row + 2,
        column=1,
        value="Wurfquote = Tore ÷ Würfe. Paradenquote = Paraden ÷ (Paraden + Gegentore). "
        "Leere Zellen bedeuten: in diesem Spiel keine Aktion erfasst.",
    ).font = NOTE_FONT

    set_widths(
        ws,
        {
            "A": 6,
            "B": 24,
            "C": 8,
            "D": 8,
            "E": 11,
            "F": 10,
            "G": 9,
            "H": 10,
            "I": 9,
            "J": 13,
            "K": 12,
            "L": 7,
            "M": 9,
            "N": 11,
            "O": 13,
        },
    )
    ws.freeze_panes = "C6"
    print_setup(ws, landscape=True, repeat_row="5:5")


def build_saison(ws) -> None:
    title_block(
        ws,
        "Saison",
        "Oben die Spiele der Saison eintragen (die Bezeichnung muss mit der Spalte „Spiel“ im Protokoll "
        "übereinstimmen). Tore und Gegentore rechnen sich daraus automatisch, ebenso die Saisonwerte weiter unten.",
    )

    game_header = 3
    style_header(ws, game_header, ["Spiel", "Datum", "Gegner", "Eigene Tore", "Gegentore"])

    first_game = game_header + 1
    last_game = first_game + 29  # 30 Spiele

    for row in range(first_game, last_game + 1):
        is_example = row == first_game
        fill = EXAMPLE_FILL if is_example else INPUT_FILL
        if is_example:
            ws.cell(row=row, column=1, value="Sp1")
            ws.cell(row=row, column=2, value="12.09.2026")
            ws.cell(row=row, column=3, value="TSV Musterstadt")

        ws.cell(
            row=row,
            column=4,
            value=(
                f'=IF($A{row}="","",'
                f'COUNTIFS({P_SPIEL},$A{row},{P_AKTION},"Tor")'
                f'+COUNTIFS({P_SPIEL},$A{row},{P_AKTION},"7m Tor"))'
            ),
        )
        ws.cell(
            row=row,
            column=5,
            value=(
                f'=IF($A{row}="","",COUNTIFS({P_SPIEL},$A{row},{P_AKTION},"Gegentor"))'
            ),
        )

        for column in range(1, 6):
            cell = ws.cell(row=row, column=column)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if column in (1, 4, 5) else "left")
            if column <= 3:
                cell.fill = fill

    # --- Saisonwerte pro Spieler ------------------------------------------
    season_header = last_game + 3
    ws.cell(row=season_header - 1, column=1, value="Saisonwerte pro Spieler").font = TITLE_FONT

    headers = [
        "Nr.",
        "Name",
        "Spiele",
        "Würfe",
        "Tore (Feld)",
        "Wurfquote",
        "7m Würfe",
        "7m Tore",
        "Tore gesamt",
        "Assists",
        "Techn. Fehler",
        "Zeitstrafen",
        "±",
        "Paraden",
        "Gegentore",
        "Paradenquote",
    ]
    style_header(ws, season_header, headers)

    first = season_header + 1
    last = first + (KADER_LAST - KADER_FIRST)
    game_list = f"$A${first_game}:$A${last_game}"

    for offset, row in enumerate(range(first, last + 1)):
        kader_row = KADER_FIRST + offset
        nr = f"$A{row}"
        guard = f'IF($A{row}="","",'

        tore = counts("Tor", nr, None)
        fehl = counts("Fehlwurf", nr, None)
        sieben_tor = counts("7m Tor", nr, None)
        sieben_fehl = counts("7m Fehlwurf", nr, None)

        formulas = {
            1: f'=IF(Kader!$A{kader_row}="","",Kader!$A{kader_row})',
            2: f'=IF(Kader!$B{kader_row}="","",Kader!$B{kader_row})',
            # Spiele mit mindestens einer erfassten Aktion.
            3: (
                f"={guard}SUMPRODUCT(({game_list}<>\"\")*"
                f"(COUNTIFS({P_SPIEL},{game_list},{P_NR},{nr})>0)))"
            ),
            4: f"={guard}{tore}+{fehl})",
            5: f"={guard}{tore})",
            6: f'={guard}IFERROR($E{row}/$D{row},""))',
            7: f"={guard}{sieben_tor}+{sieben_fehl})",
            8: f"={guard}{sieben_tor})",
            9: f"={guard}$E{row}+$H{row})",
            10: f'={guard}{counts("Assist", nr, None)})',
            11: f'={guard}{counts("Techn. Fehler", nr, None)})',
            12: f'={guard}{counts("2 Minuten", nr, None)})',
            13: f"={guard}($E{row}+$H{row})-($D{row}-$E{row})-($G{row}-$H{row})-$K{row})",
            14: f'={guard}{counts("Parade", nr, None)})',
            15: f'={guard}{counts("Gegentor", nr, None)})',
            16: f'={guard}IFERROR($N{row}/($N{row}+$O{row}),""))',
        }

        for column, formula in formulas.items():
            cell = ws.cell(row=row, column=column, value=formula)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if column != 2 else "left")
            if column in (6, 16):
                cell.number_format = PERCENT

    total_row = last + 1
    ws.cell(row=total_row, column=2, value="Mannschaft gesamt")
    for column in (4, 5, 7, 8, 9, 10, 11, 12, 13, 14, 15):
        letter = get_column_letter(column)
        ws.cell(row=total_row, column=column, value=f"=SUM({letter}{first}:{letter}{last})")
    ws.cell(row=total_row, column=6, value=f'=IFERROR($E{total_row}/$D{total_row},"")')
    ws.cell(
        row=total_row,
        column=16,
        value=f'=IFERROR($N{total_row}/($N{total_row}+$O{total_row}),"")',
    )
    for column in range(1, 17):
        cell = ws.cell(row=total_row, column=column)
        cell.font = BOLD_FONT
        cell.fill = TOTAL_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center" if column != 2 else "left")
        if column in (6, 16):
            cell.number_format = PERCENT

    ws.cell(
        row=total_row + 2,
        column=1,
        value="„Spiele“ zählt alle Partien, in denen für die Nummer mindestens eine Aktion erfasst wurde. "
        "± = Tore (inkl. 7 m) − Fehlwürfe (inkl. 7 m) − technische Fehler.",
    ).font = NOTE_FONT

    set_widths(
        ws,
        {
            "A": 6,
            "B": 24,
            "C": 8,
            "D": 8,
            "E": 12,
            "F": 11,
            "G": 10,
            "H": 9,
            "I": 12,
            "J": 9,
            "K": 13,
            "L": 12,
            "M": 7,
            "N": 9,
            "O": 11,
            "P": 13,
        },
    )
    print_setup(ws, landscape=True)

    return first_game, last_game


def main() -> None:
    workbook = Workbook()

    kader = workbook.active
    kader.title = "Kader"
    protokoll = workbook.create_sheet("Spielprotokoll")
    auswertung = workbook.create_sheet("Auswertung")
    saison = workbook.create_sheet("Saison")

    build_kader(kader)
    build_protokoll(protokoll)
    build_auswertung(auswertung)
    first_game, last_game = build_saison(saison)

    # Benannter Bereich für die Spiel-Auswahlliste auf dem Auswertungsblatt.
    workbook.defined_names["Spiele"] = DefinedName(
        "Spiele", attr_text=f"Saison!$A${first_game}:$A${last_game}"
    )

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False

    # Die Datei wird ohne zwischengespeicherte Ergebnisse geschrieben (openpyxl
    # kann Formeln nicht rechnen). Mit fullCalcOnLoad rechnet jede Tabellen-App
    # beim Öffnen einmal komplett durch, sodass sofort Werte im Blatt stehen.
    workbook.calculation.fullCalcOnLoad = True

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    print(f"geschrieben: {OUTPUT}")


if __name__ == "__main__":
    main()
